import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo 
from openpyxl.utils import get_column_letter

xls = pd.ExcelFile("KAN1.xlsx")
df = pd.read_excel(xls,index_col=False, sheet_name= "Details", na_values=["NA"], engine='openpyxl')
df2 = pd.read_excel(xls,index_col=False, sheet_name = "Summary", na_values=["NA"], engine='openpyxl')


def CollumnDeletion(x):
    x = x.drop(columns=["F"])
    return x

df = CollumnDeletion(df)

def highlight_col(row):
    if row["Echo"] == "YES":
        return ["background-color: #ffc7ce"] * len(row)
    else:
        return ["background-color: #a5a5a5"] * len(row)

df = df.style.apply(highlight_col, axis=1)
with pd.ExcelWriter("raport_charlie.xlsx") as writer:
    df2.to_excel(writer, sheet_name = "Summary", index=False)
    df.to_excel(writer, sheet_name = "Details", index=False)
    

filename = "raport_charlie.xlsx"
wb = load_workbook(filename)

ws1 = wb.worksheets[1]
tab1 = Table(displayName="Summary", ref="A1:"+ get_column_letter(ws1.max_column) + str(ws1.max_row))

ws = wb.worksheets[0]
tab = Table(displayName="Details", ref="A1:"+ get_column_letter(ws.max_column) + str(ws.max_row))

style = TableStyleInfo(
    name="TableStyleMedium1",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
tab.tableStyleInfo = style
tab1.tableStyleInfo = style
ws1.add_table(tab1)
ws.add_table(tab)

wb.save('table_example.xlsx')

#By Patryk Zachnik